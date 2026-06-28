"""Branded XLSX report builder (PRD §5.6).

Produces a workbook with three sheets:
  * Summary        — run metadata + brand header/logo
  * Results (full) — one row per SERP result
  * Brand Positions— filtered rows where the brand domain matched
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

RESULT_COLUMNS = [
    "Run Date",
    "Query",
    "Template",
    "Coin (fa_name)",
    "Coin (en_name)",
    "Coin (symbol)",
    "Position",
    "Page",
    "Domain",
    "Title",
    "URL",
    "Snippet",
    "Is Brand",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BRAND_FILL = PatternFill("solid", fgColor="FFF3CD")


def _write_brand_header(ws, config: dict[str, Any], run_date: str) -> None:
    """Rows 1–3: logo (A1), brand name, config snapshot (PRD §5.6.3)."""
    brand_name = config.get("brand_name", "") or "SERP Position Tracker"
    logo_path = config.get("brand_logo_path", "")

    ws["A1"] = brand_name
    ws["A1"].font = Font(bold=True, size=16, color="1F3B4D")
    ws["A2"] = (
        f"Region: {config.get('region', '')}  |  Language: {config.get('language', '')}"
        f"  |  Device: {config.get('device', '')}  |  Pages: {config.get('num_pages', '')}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A3"] = f"Run Date: {run_date}"
    ws["A3"].font = Font(size=10, color="555555")

    if logo_path and Path(logo_path).exists():
        try:
            img = XLImage(logo_path)
            img.height = min(img.height, 60)
            img.width = min(img.width, 180)
            ws.add_image(img, "H1")
        except Exception:
            pass  # logo is best-effort; never block the report


def _style_table_header(ws, header_row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, df: pd.DataFrame, start_col: int = 1) -> None:
    for i, col in enumerate(df.columns, start=start_col):
        width = max(len(str(col)), *(df[col].astype(str).map(len).tolist() or [0]))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)


def _append_dataframe(ws, df: pd.DataFrame, start_row: int) -> int:
    header_row = start_row
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row + r_idx, column=c_idx, value=value)
    _style_table_header(ws, header_row, len(df.columns))
    return header_row


def build_report(
    rows: list[dict[str, Any]],
    config: dict[str, Any],
    run_date: str,
    query_count: int,
    output_path: str | Path,
    error_count: int = 0,
) -> str:
    """Write the branded XLSX workbook and return its path."""
    df = pd.DataFrame(rows, columns=[c for c in RESULT_COLUMNS if c != "Run Date"])
    df.insert(0, "Run Date", run_date)

    wb = Workbook()

    # --- Sheet 1: Summary ------------------------------------------------- #
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_brand_header(ws_sum, config, run_date)
    summary = pd.DataFrame(
        [
            ("Brand", config.get("brand_name", "")),
            ("Run Date", run_date),
            ("Total Queries", query_count),
            ("Total Result Rows", len(df)),
            ("Brand Matches", int(df["Is Brand"].sum()) if not df.empty else 0),
            ("Errors", error_count),
            ("Region (gl)", config.get("region", "")),
            ("Language (hl)", config.get("language", "")),
            ("Device", config.get("device", "")),
            ("Pages Retrieved", config.get("num_pages", "")),
        ],
        columns=["Metric", "Value"],
    )
    _append_dataframe(ws_sum, summary, start_row=5)
    _autosize(ws_sum, summary)

    # --- Sheet 2: Results (full) ----------------------------------------- #
    ws_res = wb.create_sheet("Results (full)")
    _write_brand_header(ws_res, config, run_date)
    _append_dataframe(ws_res, df, start_row=5)
    _autosize(ws_res, df)
    # Highlight brand rows.
    brand_col = list(df.columns).index("Is Brand") + 1
    for r in range(6, 6 + len(df)):
        if ws_res.cell(row=r, column=brand_col).value is True:
            for c in range(1, len(df.columns) + 1):
                ws_res.cell(row=r, column=c).fill = _BRAND_FILL

    # --- Sheet 3: Brand Positions ---------------------------------------- #
    ws_brand = wb.create_sheet("Brand Positions")
    _write_brand_header(ws_brand, config, run_date)
    brand_df = df[df["Is Brand"] == True].reset_index(drop=True)  # noqa: E712
    if brand_df.empty:
        ws_brand["A5"] = "No brand domains matched in this run."
        ws_brand["A5"].font = Font(italic=True, color="999999")
    else:
        _append_dataframe(ws_brand, brand_df, start_row=5)
        _autosize(ws_brand, brand_df)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(output_path)
